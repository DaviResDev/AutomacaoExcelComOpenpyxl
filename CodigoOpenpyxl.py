- Feito para estudo pessoal, foi instalado o Openpyxl no terminal usando "pip install"; e este é o código base para estudar e testar, de forma básica, a estrutura da biblioteca.
-A Base de dados está anexada no mesmo repositório.

from openpyxl import load_workbook

arquivo = load_workbook("Alunos.xlsx")

# ver as abas
print(arquivo.sheetnames)

# pegar a aba ativa
aba_atual = arquivo.active
print(aba_atual)

# selecionar uma aba específica
aba_alunos = arquivo["Planilha1"]
print(aba_alunos)

# selecionar células
valor_a1 = aba_alunos["A1"].value
valor_b1 = aba_alunos.cell(row=1, column=2).value
print(valor_b1)

aba_alunos.cell(row=1, column=2).value = "Prova 1"

arquivo.save("Alunos2.xlsx")

# ultima linha
print(aba_alunos.max_row)
print(len(aba_alunos["A"]))
